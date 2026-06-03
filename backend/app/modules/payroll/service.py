"""Payroll use-cases: component/formula config, context building, run
calculation with reproducible snapshots, locking/cancel, and Excel import.

The engine maths live in :mod:`app.modules.payroll.engine`; this layer wires it
to the database, encryption and audit trail.
"""

from __future__ import annotations

import calendar
import io
import re
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any

from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.audit.recorder import record
from app.core.encryption import decrypt_decimal, encrypt_decimal
from app.core.exceptions import ConflictError, NotFoundError, ValidationError
from app.core.logging import get_logger
from app.core.rbac import CurrentUser
from app.modules.attendance.models import AttendanceMonthly
from app.modules.employee.models import Employee
from app.modules.payroll.engine import build_eval_order, evaluate
from app.modules.payroll.models import (
    NET_VAR,
    PayrollOverride,
    PayrollPeriod,
    PayrollRun,
    PayrollRunItem,
    PeriodStatus,
    RunStatus,
    SalaryComponent,
    ValueType,
)
from app.modules.payroll.repository import (
    ComponentRepository,
    InputValueRepository,
    OverrideRepository,
    PeriodRepository,
    RunItemRepository,
    RunRepository,
)

logger = get_logger("payroll.service")

# Variables always present in an employee's context (no formula needed).
BASE_VARS: frozenset[str] = frozenset(
    {"luong_cung", "company_salary", "cong_chuan", "cong_thuc_te", "ot_gio"}
)


def _slugify(value: str) -> str:
    slug = re.sub(r"[^0-9a-zA-Z]+", "_", value.strip().lower()).strip("_")
    return slug or "var"


def _period_bounds(code: str) -> tuple[date, date]:
    year, month = (int(p) for p in code.split("-"))
    last = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last)


class PayrollService:
    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.components = ComponentRepository(session)
        self.periods = PeriodRepository(session)
        self.runs = RunRepository(session)
        self.items = RunItemRepository(session)
        self.inputs = InputValueRepository(session)
        self.overrides = OverrideRepository(session)

    # ----------------------------------------------------------------- #
    # Component / formula config                                        #
    # ----------------------------------------------------------------- #
    async def _base_var_names(self) -> set[str]:
        """Base vars + every active INPUT/FIXED component var_code."""
        names = set(BASE_VARS)
        for comp in await self.components.active():
            if comp.value_type in (ValueType.INPUT, ValueType.FIXED):
                names.add(comp.var_code)
        return names

    async def _validate_formula_graph(self, extra: tuple[str, str] | None = None) -> None:
        """Validate the formula DAG (cycles + undefined vars).

        ``extra`` is a (target, expression) being added/edited; included so the
        graph is validated *before* persisting (fail fast at save time).
        """
        formulas = [
            (c.var_code, c.expression or "") for c in await self.components.active_formulas()
        ]
        if extra is not None:
            formulas = [f for f in formulas if f[0] != extra[0]] + [extra]
        build_eval_order(formulas, await self._base_var_names())

    async def create_component(
        self,
        *,
        code: str,
        name: str,
        value_type: str,
        var_code: str | None = None,
        default_value: Decimal | None = None,
        expression: str | None = None,
    ) -> SalaryComponent:
        if value_type not in ValueType.ALL:
            raise ValidationError("Loại khoản lương không hợp lệ")
        if await self.components.get_by_code(code):
            raise ConflictError("Mã khoản lương đã tồn tại")

        var_code = var_code or _slugify(name)
        if await self.components.get_by_var_code(var_code):
            raise ConflictError(f"var_code '{var_code}' đã tồn tại")

        if value_type == ValueType.FORMULA and not expression:
            raise ValidationError("Khoản lương FORMULA cần biểu thức")
        if value_type == ValueType.FIXED and default_value is None:
            raise ValidationError("Khoản lương FIXED cần giá trị mặc định")

        # Validate the DAG including this formula *before* saving.
        if value_type == ValueType.FORMULA:
            await self._validate_formula_graph(extra=(var_code, expression or ""))

        component = SalaryComponent(
            code=code,
            var_code=var_code,
            name=name,
            value_type=value_type,
            default_value=default_value,
            expression=expression,
        )
        return await self.components.add(component)

    # ----------------------------------------------------------------- #
    # Periods                                                           #
    # ----------------------------------------------------------------- #
    async def get_or_create_period(
        self, code: str, *, standard_days: Decimal | None = None
    ) -> PayrollPeriod:
        period = await self.periods.get_by_code(code)
        if period is None:
            period = PayrollPeriod(code=code, standard_days=standard_days or Decimal("22.00"))
            await self.periods.add(period)
        elif standard_days is not None:
            period.standard_days = standard_days
            await self.session.flush()
        return period

    # ----------------------------------------------------------------- #
    # Context building                                                  #
    # ----------------------------------------------------------------- #
    async def build_context(self, employee: Employee, period: PayrollPeriod) -> dict[str, float]:
        """Assemble the variable context for one employee (docs/03b §3.4)."""
        start, end = _period_bounds(period.code)
        base_salary = (
            decrypt_decimal(employee.enc_base_salary) if employee.enc_base_salary else None
        ) or Decimal("0")
        ctx: dict[str, float] = {
            "luong_cung": float(base_salary),
            "company_salary": float(base_salary),
        }

        monthly = (
            await self.session.execute(
                select(AttendanceMonthly).where(
                    AttendanceMonthly.employee_id == employee.id,
                    AttendanceMonthly.period == period.code,
                )
            )
        ).scalar_one_or_none()
        ctx["cong_chuan"] = float(monthly.standard_days if monthly else period.standard_days)
        ctx["cong_thuc_te"] = float(monthly.actual_days) if monthly else 0.0
        ctx["ot_gio"] = float(monthly.ot_hours) if monthly else 0.0

        # Component vars (INPUT coalesced to 0, FIXED -> default).
        input_map = await self.inputs.map_for(period.id, employee.id)
        components = await self.components.for_employee(
            department_id=employee.department_id,
            position_id=employee.position_id,
            employee_id=employee.id,
            start=start,
            end=end,
        )
        for comp in components:
            if comp.value_type == ValueType.INPUT:
                ctx[comp.var_code] = float(input_map.get(comp.id, Decimal("0")))
            elif comp.value_type == ValueType.FIXED:
                ctx[comp.var_code] = float(comp.default_value or Decimal("0"))
            # FORMULA components are evaluated, not seeded.

        # Overrides (e.g. maternity: company_salary=0) merged last.
        override = await self.overrides.get_for(period.id, employee.id)
        if override is not None:
            for key, value in override.data.items():
                if isinstance(value, (int, float, bool)):
                    ctx[key] = value if isinstance(value, bool) else float(value)
        return ctx

    # ----------------------------------------------------------------- #
    # Run lifecycle                                                     #
    # ----------------------------------------------------------------- #
    async def create_run(
        self, period_code: str, actor: CurrentUser, *, ip: str | None = None
    ) -> PayrollRun:
        period = await self.get_or_create_period(period_code)
        if await self.runs.active_for_period(period.id):
            raise ConflictError("Kỳ lương đã có một bảng tính đang hoạt động")

        # Snapshot active formulas (validates the DAG too) for reproducibility.
        await self._validate_formula_graph()
        snapshot = [
            {"target_var": c.var_code, "expression": c.expression or ""}
            for c in await self.components.active_formulas()
        ]
        run = PayrollRun(
            period_id=period.id,
            status=RunStatus.DRAFT,
            formula_snapshot=snapshot,
            created_by=actor.id,
        )
        await self.runs.add(run)
        await record(
            self.session,
            actor_id=actor.id,
            action="CREATE",
            entity="payroll_runs",
            entity_id=run.id,
            new={"period": period_code},
            ip=ip,
        )
        return run

    async def calculate_run(self, run_id: int, *, employee_ids: list[int] | None = None) -> int:
        """Compute payroll for a DRAFT run from its frozen formula snapshot.

        Idempotent: existing items are cleared and recomputed. Returns the
        number of employees calculated.
        """
        run = await self.runs.get(run_id)
        if run is None:
            raise NotFoundError("Không tìm thấy bảng tính lương")
        if run.status != RunStatus.DRAFT:
            raise ConflictError("Chỉ được tính lại bảng tính ở trạng thái DRAFT")
        period = await self.periods.get(run.period_id)
        assert period is not None

        snapshot = run.formula_snapshot or []
        formula_by_target = {f["target_var"]: f["expression"] for f in snapshot}
        eval_order = build_eval_order(
            [(f["target_var"], f["expression"]) for f in snapshot],
            await self._base_var_names(),
        )

        employees = await self._employees(employee_ids)
        # Clear only what we're about to recompute, so parallel Celery chunks
        # (each a disjoint employee slice) don't wipe each other's items.
        if employee_ids:
            await self.items.delete_for_employees(run_id, employee_ids)
        else:
            await self.items.delete_for_run(run_id)

        for employee in employees:
            ctx = await self.build_context(employee, period)
            full = evaluate(formula_by_target, eval_order, ctx)
            net = full.get(NET_VAR)
            input_snapshot = {k: v for k, v in full.items() if k not in formula_by_target}
            self.session.add(
                PayrollRunItem(
                    run_id=run_id,
                    employee_id=employee.id,
                    input_snapshot=input_snapshot,
                    result=full,
                    enc_net_amount=encrypt_decimal(net) if net is not None else None,
                )
            )
        await self.session.flush()
        logger.info("payroll_run_calculated", run_id=run_id, employees=len(employees))
        return len(employees)

    async def _employees(self, employee_ids: list[int] | None) -> list[Employee]:
        stmt = select(Employee).where(Employee.is_deleted.is_(False))
        if employee_ids:
            stmt = stmt.where(Employee.id.in_(employee_ids))
        return list((await self.session.execute(stmt)).scalars().all())

    async def lock_run(
        self, run_id: int, actor: CurrentUser, *, ip: str | None = None
    ) -> PayrollRun:
        run = await self.runs.get_locked(run_id)
        if run is None:
            raise NotFoundError("Không tìm thấy bảng tính lương")
        if run.status != RunStatus.DRAFT:
            raise ConflictError("Chỉ khóa được bảng tính DRAFT")
        period = await self.periods.get(run.period_id)
        assert period is not None

        locked_at = datetime.now(UTC)
        run.status = RunStatus.LOCKED
        run.locked_at = locked_at
        period.status = PeriodStatus.LOCKED
        # Freeze attendance for the period so it cannot drift after payroll.
        await self.session.execute(
            update(AttendanceMonthly)
            .where(AttendanceMonthly.period == period.code)
            .values(locked=True)
        )
        await self.session.flush()
        await record(
            self.session,
            actor_id=actor.id,
            action="LOCK",
            entity="payroll_runs",
            entity_id=run.id,
            new={"locked_at": locked_at.isoformat()},
            ip=ip,
        )
        return run

    async def cancel_run(
        self, run_id: int, actor: CurrentUser, *, ip: str | None = None
    ) -> PayrollRun:
        run = await self.runs.get_locked(run_id)
        if run is None:
            raise NotFoundError("Không tìm thấy bảng tính lương")
        if run.status not in RunStatus.ACTIVE:
            raise ConflictError("Bảng tính không ở trạng thái có thể hủy")
        period = await self.periods.get(run.period_id)
        assert period is not None

        run.status = RunStatus.CANCELLED
        period.status = PeriodStatus.OPEN
        await self.session.execute(
            update(AttendanceMonthly)
            .where(AttendanceMonthly.period == period.code)
            .values(locked=False)
        )
        # Items kept as immutable history but flagged cancelled.
        await self.session.execute(
            update(PayrollRunItem).where(PayrollRunItem.run_id == run_id).values(status="CANCELLED")
        )
        await self.session.flush()
        await record(
            self.session,
            actor_id=actor.id,
            action="CANCEL",
            entity="payroll_runs",
            entity_id=run.id,
            ip=ip,
        )
        return run

    # ----------------------------------------------------------------- #
    # Overrides                                                         #
    # ----------------------------------------------------------------- #
    async def set_override(
        self, *, period_code: str, employee_id: int, data: dict[str, Any]
    ) -> None:
        period = await self.get_or_create_period(period_code)
        existing = await self.overrides.get_for(period.id, employee_id)
        if existing is None:
            self.session.add(
                PayrollOverride(period_id=period.id, employee_id=employee_id, data=data)
            )
        else:
            existing.data = data
        await self.session.flush()

    # ----------------------------------------------------------------- #
    # Excel import / template                                           #
    # ----------------------------------------------------------------- #
    async def _active_input_components(self) -> list[SalaryComponent]:
        return [c for c in await self.components.active() if c.value_type == ValueType.INPUT]

    async def generate_template(self, period_code: str) -> bytes:
        """Build an .xlsx template: columns = employee_code + INPUT var_codes."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = period_code
        headers = ["employee_code"] + [c.var_code for c in await self._active_input_components()]
        ws.append(headers)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def import_input(self, *, period_code: str, file_bytes: bytes) -> dict[str, Any]:
        """Import INPUT values from an .xlsx. Returns an ok/errors report."""
        from openpyxl import load_workbook

        period = await self.get_or_create_period(period_code)
        components = {c.var_code: c for c in await self._active_input_components()}

        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header or header[0] != "employee_code":
            raise ValidationError("File không hợp lệ: thiếu cột 'employee_code'")
        var_cols = list(header[1:])

        ok = 0
        errors: list[dict[str, Any]] = []
        for line_no, row in enumerate(rows, start=2):
            if row is None or row[0] in (None, ""):
                continue
            emp_code = str(row[0])
            employee = await self._employee_by_code(emp_code)
            if employee is None:
                errors.append({"row": line_no, "error": f"Không có NV '{emp_code}'"})
                continue
            row_ok = True
            for idx, var_code in enumerate(var_cols, start=1):
                raw = row[idx] if idx < len(row) else None
                if raw in (None, ""):
                    continue
                comp = components.get(str(var_code))
                if comp is None:
                    errors.append({"row": line_no, "error": f"var_code lạ '{var_code}'"})
                    row_ok = False
                    continue
                try:
                    value = Decimal(str(raw))
                except Exception:  # noqa: BLE001
                    errors.append({"row": line_no, "error": f"'{var_code}' không phải số"})
                    row_ok = False
                    continue
                if value < 0:
                    errors.append({"row": line_no, "error": f"'{var_code}' âm"})
                    row_ok = False
                    continue
                await self.inputs.upsert(
                    period_id=period.id,
                    employee_id=employee.id,
                    component_id=comp.id,
                    value=value,
                )
            if row_ok:
                ok += 1
        return {"ok": ok, "errors": errors}

    async def _employee_by_code(self, code: str) -> Employee | None:
        return (
            await self.session.execute(
                select(Employee).where(
                    Employee.employee_code == code, Employee.is_deleted.is_(False)
                )
            )
        ).scalar_one_or_none()
